#!/usr/bin/env python
# -*- encoding:utf-8 -*-
"""
@Author : 卢晋
@File   : util.py
@Time   : 2023年5月15日
@Desc   : 工具类
"""

import pdb
import openpyxl
import os
import re
import sys
from decimal import Decimal

MIN_SIZE = -sys.maxsize - 1

def exportExcel(table, filename, progress=None, taskbar_progress=None):
    wb = openpyxl.Workbook()
    ws = wb.active
    
    max_row = table.rowCount()
    max_col = table.columnCount()
    total = max_col * max_row

    if progress:
        progress.setRange(0, total)
        progress.setValue(0)
    if sys.platform == 'win32' and taskbar_progress != None:
        taskbar_progress.setRange(0, total)
        taskbar_progress.setValue(0)

    # 测试值
    done = 0
    for col in range(0, max_col):
        col_letter = ws.cell(1, col + 1).column_letter
        max_length = ws.column_dimensions[col_letter].width
        it = iter(range(1, max_row))
        for row in it:
            if progress != None and progress.wasCanceled():
                raise Exception('user canceled')
            item = table.item(row, col)
            if not item:
                done += 1
                if progress != None:
                    progress.setValue(done)
                if sys.platform == 'win32' and taskbar_progress != None:
                    taskbar_progress.setValue(done)
                continue
            span = table.rowSpan(row, col)
            done += span
            cell = ws.cell(row + 1, col + 1)
            text = item.text().strip()
            if text[0] == '-':
                text = text[1:]
            if span > 1:
                cell.value = item.text().strip()
            elif text.find('.') > -1:
                integer, decimal = text.split('.')
                if integer.isdigit() and decimal.isdigit():
                    cell.value = float(item.text().strip())
                else:
                    cell.value = item.text().strip()
            elif text.isdigit():
                cell.value = int(item.text().strip())
            else:
                cell.value = item.text().strip()
            rgb = item.background().color().rgb() % (256 * 256 * 256)
            if rgb != 0:
                cell.fill = openpyxl.styles.PatternFill(patternType='solid',fgColor=hex(rgb)[2:])
            max_length = max(max_length, len(str(cell.value)))
            if span > 1:
                ws.merge_cells(start_row = row + 1, start_column = col + 1, end_row = row + span, end_column = col + 1)
                cell.alignment = openpyxl.styles.alignment.Alignment(horizontal='center', vertical='top')
                for i in range(span - 1):
                    next(it)
            else:
                cell.alignment = openpyxl.styles.alignment.Alignment(horizontal='center', vertical='center')
            if progress != None:
                progress.setValue(done)
            if sys.platform == 'win32' and taskbar_progress != None:
                taskbar_progress.setValue(done)
        adjusted_width = (max_length + 2) * 1.2
        if adjusted_width > max_length:
            # col_letter = [x[0].column_letter for x in ws.columns][col]
            ws.column_dimensions[col_letter].width = adjusted_width

    # # 芯片ID
    # it = iter(range(max_row))
    # for row in it:
    #     span = table.rowSpan(row, 0)
    #     item = table.item(row, 0)
    #     if not item:
    #         continue
    #     cell = ws.cell(row + 1, 1)
    #     cell.value = item.text()
    #     cell.alignment = openpyxl.styles.alignment.Alignment(horizontal='center', vertical='top')
    #     ws.merge_cells(start_row = row + 1, start_column = 1, end_row = row + span, end_column = 1)
    #     if span > 1:
    #         for i in range(span - 1):
    #             next(it)

    done = total - max_col
    if progress != None:
        progress.setValue(done)
    if sys.platform == 'win32' and taskbar_progress != None:
        taskbar_progress.setValue(done)
    # 表头
    for row in range(1):
        it = iter(range(max_col))
        for col in it:
            item = table.item(row, col)
            if not item:
                done += 1
                if progress != None:
                    progress.setValue(done)
                if sys.platform == 'win32' and taskbar_progress != None:
                    taskbar_progress.setValue(done)
                continue
            span = table.columnSpan(row, col)
            done += span
            cell = ws.cell(row + 1, col + 1)
            cell.value = item.text()
            cell.alignment = openpyxl.styles.alignment.Alignment(horizontal='center', vertical='center')
            ws.merge_cells(start_row = row + 1, start_column = col + 1, end_row = row + 1, end_column = col + span)
            if span > 1:
                for i in range(span - 1):
                    next(it)
            if progress != None:
                progress.setValue(done)
            if sys.platform == 'win32' and taskbar_progress != None:
                taskbar_progress.setValue(done)

    if progress != None:
        progress.reset()
    if sys.platform == 'win32' and taskbar_progress != None:
        taskbar_progress.resume()
        taskbar_progress.reset()
    wb.save(filename)

def isnumber(text):
    if not text and len(text) == 0:
        return False
    if text[0] == '-':
        text = text[1:]
    vals = text.split('.')
    if len(vals) > 2:
        return False
    for val in vals:
        if not val.isdigit():
            return False
    return True
